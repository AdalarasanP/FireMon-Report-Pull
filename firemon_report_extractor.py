#!/usr/bin/env python3
"""
FireMon Report Schedule Extractor
Fetches report schedules from FireMon API and exports to Excel

This script extracts report schedules from FireMon Security Manager API and generates
comprehensive Excel reports with multiple views including:
- Detailed schedule listings
- Matrix views by tower/department
- Time slot analysis
- Conflict detection and highlighting
"""

import argparse
import json
import os
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import urllib3


def load_dotenv(path: str = '.env') -> dict:
    """Load environment variables from .env file"""
    env = {}
    p = Path(path)
    if not p.exists():
        return env
    with open(p, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                env[key.strip()] = val.strip().strip('"').strip("'")
    return env


def fetch_reports(base_url: str, username: str, password: str, verify_ssl: bool = True):
    """Fetch all reports from FireMon API with pagination"""
    all_results = []
    page_size = 1000
    offset = 0
    auth = HTTPBasicAuth(username, password)
    
    while True:
        url = f"{base_url}/report?pageSize={page_size}&offset={offset}"
        print(f"Fetching from: {url}")
        
        try:
            resp = requests.get(url, auth=auth, verify=verify_ssl, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            
            results = data.get('results', [])
            if not results:
                break
                
            all_results.extend(results)
            
            total = data.get('total', 0)
            if len(all_results) >= total:
                break
                
            offset += page_size
            
        except requests.exceptions.RequestException as e:
            print(f"Error: {e}")
            raise  # Raise the exception to trigger fallback to cache
    
    if not all_results:
        raise Exception("No reports fetched from API")
    
    return all_results


def fetch_device_groups(base_url: str, username: str, password: str, verify_ssl: bool = True):
    """Fetch all device groups from FireMon API"""
    all_results = []
    page_size = 1000
    offset = 0
    auth = HTTPBasicAuth(username, password)
    
    while True:
        url = f"{base_url}/devicegroup?pageSize={page_size}&offset={offset}"
        print(f"Fetching device groups from: {url}")
        
        try:
            resp = requests.get(url, auth=auth, verify=verify_ssl, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            
            results = data.get('results', [])
            if not results:
                break
                
            all_results.extend(results)
            
            total = data.get('total', 0)
            if len(all_results) >= total:
                break
                
            offset += page_size
            
        except requests.exceptions.RequestException as e:
            print(f"Error: {e}")
            raise  # Raise the exception to trigger fallback to cache
    
    if not all_results:
        raise Exception("No device groups fetched from API")
    
    return all_results


def adjust_datetime_fields(obj, offset_hours=-5):
    """
    Adjust datetime fields by offset_hours
    
    Args:
        obj: Dictionary or list containing datetime strings
        offset_hours: Hours to adjust (default: -5 for EST)
    """
    modified_count = 0
    
    if isinstance(obj, dict):
        for key, value in obj.items():
            if isinstance(value, str) and 'T' in value and ('Z' in value or '+' in value or '-' in value[-6:]):
                try:
                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    adjusted = dt + timedelta(hours=offset_hours)
                    obj[key] = adjusted.strftime('%Y-%m-%dT%H:%M:%S')
                    modified_count += 1
                except (ValueError, AttributeError):
                    pass
            elif isinstance(value, (dict, list)):
                modified_count += adjust_datetime_fields(value, offset_hours)
    elif isinstance(obj, list):
        for item in obj:
            modified_count += adjust_datetime_fields(item, offset_hours)
    
    return modified_count


def main():
    parser = argparse.ArgumentParser(description='Extract FireMon report schedules to Excel')
    parser.add_argument('--prefix', default='firemon-report', help='Output file prefix')
    parser.add_argument('--timezone-offset', type=int, default=-5, help='Timezone offset from UTC (default: -5 for EST)')
    
    args = parser.parse_args()
    
    # Load environment
    env = load_dotenv('.env')
    base_url = env.get('FIREMON_BASE_URL')
    username = env.get('FIREMON_USER')
    password = env.get('FIREMON_PASS')
    verify_ssl = env.get('FIREMON_VERIFY_SSL', 'false').lower() == 'true'
    
    if not base_url:
        print("Error: FIREMON_BASE_URL must be set in .env file")
        print("Example: FIREMON_BASE_URL=https://your-firemon-server.com/securitymanager/api/domain/1")
        return
    
    if not username or not password:
        print("Error: FIREMON_USER and FIREMON_PASS must be set in .env file")
        return
    
    # Disable SSL warnings if SSL verification is disabled
    if not verify_ssl:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    # Try to fetch real-time data from FireMon API, fall back to cache if auth fails
    output_json = f'{args.prefix}.json'
    device_groups_json = f'{args.prefix}-devicegroups.json'
    
    try:
        print("Fetching real-time data from FireMon...")
        print("Fetching reports...")
        reports = fetch_reports(base_url, username, password, verify_ssl)
        print(f"Fetched {len(reports)} reports")
        
        print("\nFetching device groups...")
        device_groups = fetch_device_groups(base_url, username, password, verify_ssl)
        print(f"Fetched {len(device_groups)} device groups")
        
        # Adjust datetime fields
        modified = adjust_datetime_fields(reports, offset_hours=args.timezone_offset)
        print(f"Adjusted {modified} datetime fields")
        
        # Save reports JSON (for reference)
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(reports, f, indent=2)
        print(f"Saved reports to {output_json}")
        
        # Save device groups JSON (for reference)
        with open(device_groups_json, 'w', encoding='utf-8') as f:
            json.dump(device_groups, f, indent=2)
        print(f"Saved device groups to {device_groups_json}")
    except Exception as e:
        # Fall back to cached data if available
        print(f"\nWarning: Could not fetch data from API: {e}")
        print("Attempting to use cached data...")
        
        if os.path.exists(output_json) and os.path.exists(device_groups_json):
            with open(output_json, 'r', encoding='utf-8') as f:
                reports = json.load(f)
            with open(device_groups_json, 'r', encoding='utf-8') as f:
                device_groups = json.load(f)
            print(f"Loaded {len(reports)} reports from cache")
            print(f"Loaded {len(device_groups)} device groups from cache")
        else:
            print("Error: No cached data available")
            return
    
    # Create device group lookup
    device_group_map = {dg.get('id'): dg.get('name', f"DG-{dg.get('id')}") for dg in device_groups}
    
    # Function to infer tower/department from device group name and report name
    # CUSTOMIZE THIS FUNCTION for your organization's naming conventions
    def infer_tower(device_group_name: str, report_name: str) -> str:
        """
        Infer tower/department from device group name and report name using keyword matching
        
        IMPORTANT: Customize this function for your organization's naming conventions!
        
        Example towers/departments:
        - DATA_CENTER: Data center devices
        - CORPORATE: Corporate network devices
        - WAN: Wide area network devices
        - CLOUD: Cloud network devices
        - BRANCH: Branch/property devices
        """
        dg_lower = device_group_name.lower()
        report_lower = report_name.lower()
        
        # Rule-based classification examples:
        # Customize these rules based on your organization's naming patterns
        
        # Example 1: Match by report name keywords
        if 'datacenter' in report_lower or 'data center' in report_lower:
            return 'DATA_CENTER'
        
        if 'cloud' in report_lower or 'cloudnetworks' in report_lower:
            return 'CLOUD'
        
        if 'corporate' in report_lower or 'enterprise' in report_lower:
            return 'CORPORATE'
        
        if 'wan' in report_lower or 'wide area' in report_lower:
            return 'WAN'
        
        if 'property' in report_lower or 'branch' in report_lower:
            return 'BRANCH'
        
        # Example 2: Match by device group name patterns
        if any(kw in dg_lower for kw in ['corp-', 'corporate', 'ent-', 'enterprise']):
            return 'CORPORATE'
        
        if any(kw in dg_lower for kw in ['wan', 'wide area']):
            return 'WAN'
        
        if any(kw in dg_lower for kw in ['dc-', 'datacenter', 'data-center']):
            return 'DATA_CENTER'
        
        if any(kw in dg_lower for kw in ['cloud', 'azure', 'aws', 'gcp']):
            return 'CLOUD'
        
        # Default fallback
        return 'OTHER'
    
    # Function to format time in different timezones
    def format_time_local(time_str: str, offset_hours: int = -5, tz_name: str = 'EST') -> str:
        """
        Format time to local timezone
        
        Args:
            time_str: ISO format time string
            offset_hours: Timezone offset from UTC (default: -5 for EST)
            tz_name: Timezone name for display (default: EST)
        """
        if not time_str or time_str == 'N/A':
            return 'N/A'
        try:
            # If time has Z suffix, it's UTC and needs conversion
            if 'Z' in time_str:
                dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                local_dt = dt + timedelta(hours=offset_hours)
                return local_dt.strftime('%I:%M %p')
            else:
                # No timezone indicator - assume already in local timezone
                dt = datetime.fromisoformat(time_str)
                return dt.strftime('%I:%M %p')
        except:
            return time_str
    
    # Function to create friendly frequency description
    def get_frequency_description(scheduled_task: dict) -> str:
        """Create human-readable frequency description"""
        if not scheduled_task.get('enabled', False):
            return 'Not Scheduled'
        
        recurrence = scheduled_task.get('recurrence', '')
        weekdays = scheduled_task.get('weekdays', [])
        start_time = scheduled_task.get('startTime', '')
        
        # Extract time from start_time
        time_part = 'N/A'
        if start_time:
            try:
                dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                time_part = dt.strftime('%I:%M %p')
            except:
                pass
        
        if recurrence == 'WEEKLY' and weekdays:
            days_str = ', '.join(weekdays)
            return f"Weekly on {days_str} at {time_part}"
        elif recurrence == 'DAILY':
            return f"Daily at {time_part}"
        elif recurrence == 'MONTHLY':
            return f"Monthly at {time_part}"
        else:
            return f"{recurrence} at {time_part}"
    
    # Extract report schedule data
    schedule_data = []
    
    for report in reports:
        extended = report.get('extendedSettings', {})
        scheduled_task = report.get('scheduledTask', {})
        
        report_name = extended.get('name', 'Unknown')
        device_group_id = extended.get('deviceGroupId')
        
        # Skip reports without device group
        if not device_group_id:
            continue
        
        # Try to get device group name from the lookup, or extract from report name
        device_group_name = device_group_map.get(device_group_id, f"DG-{device_group_id}")
        
        # If we got a generic name, try to extract from report name
        if device_group_name.startswith('DG-'):
            # Try to parse device group name from report name
            # Common patterns: "DeviceGroupName - Category - Report..."  or "DeviceGroupName-Category-Report..."
            if ' - ' in report_name:
                parts = report_name.split(' - ')
                if len(parts) > 0 and parts[0].strip():
                    device_group_name = parts[0].strip()
        
        # Infer tower using both device group name and report name
        tower = infer_tower(device_group_name, report_name)
        
        # Schedule details
        schedule_enabled = scheduled_task.get('enabled', False)
        recurrence = scheduled_task.get('recurrence', 'N/A')
        weekdays = ', '.join(scheduled_task.get('weekdays', [])) if scheduled_task.get('weekdays') else 'N/A'
        # For monthly schedules, use 'days' field (list of day numbers)
        days = scheduled_task.get('days', [])
        day_of_month = ', '.join(str(d) for d in days) if days else 'N/A'
        
        # Format times
        start_time_local = format_time_local(scheduled_task.get('startTime', ''), args.timezone_offset)
        next_fire_local = format_time_local(scheduled_task.get('nextFireTime', ''), args.timezone_offset)
        
        # Get friendly frequency description
        frequency_desc = get_frequency_description(scheduled_task)
        
        schedule_data.append({
            'Tower': tower,
            'Device Group Name': device_group_name,
            'Device Group ID': device_group_id,
            'Report Name': report_name,
            'Report ID': report.get('id'),
            'Schedule Enabled': 'Yes' if schedule_enabled else 'No',
            'Recurrence': recurrence,
            'Weekdays': weekdays,
            'Day of Month': day_of_month,
            'Frequency': frequency_desc,
            'Start Time': start_time_local,
            'Next Fire Time': next_fire_local,
        })
    
    # Helper function to extract category from report name
    # CUSTOMIZE THIS FUNCTION for your organization's report types
    def extract_category(report_name: str, tower: str = '') -> str:
        """
        Classify report into predefined categories based on keywords in report name
        
        IMPORTANT: Customize this function for your organization's report categories!
        """
        report_lower = report_name.lower()
        
        # Standard FireMon report categories examples:
        if 'wide open' in report_lower or 'wide-open' in report_lower:
            return 'Wide-open rules'
        elif 'overly permissive' in report_lower:
            return 'Overly Permissive'
        elif 'non-standard port' in report_lower:
            return 'Non-Standard port'
        elif 'implied deny' in report_lower:
            return 'Implied Deny'
        elif 'unused rule' in report_lower:
            return 'Unused rules'
        elif 'unused object' in report_lower:
            return 'Unused object'
        elif 'shadow' in report_lower:
            return 'Shadow rules'
        elif 'compliance' in report_lower:
            return 'OS Compliance'
        elif 'change report' in report_lower or 'change tracking' in report_lower:
            return 'Change Report'
        elif 'golden config' in report_lower:
            return 'Golden Config'
        elif 'blacklist' in report_lower:
            return 'Blacklisted IP Allowed'
        else:
            return 'Miscellaneous/Adhoc'
    
    # Helper function to create time slot analysis
    def create_timeslot_analysis(data_df: pd.DataFrame) -> pd.DataFrame:
        """Create a time slot analysis showing which device groups are scheduled at each time for each category"""
        if data_df.empty:
            return pd.DataFrame()
        
        # Add category column
        data_df = data_df.copy()
        data_df['Category'] = data_df.apply(lambda row: extract_category(row['Report Name'], row.get('Tower', '')), axis=1)
        
        # Get all categories
        all_categories = sorted(data_df['Category'].unique())
        
        # Generate time slots (30-minute intervals for 24 hours)
        time_slots = []
        for hour in range(24):
            for minute in [0, 30]:
                time_slots.append(f"{hour:02d}:{minute:02d}")
        
        # Generate weekday/time rows
        weekdays = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
        rows = []
        for day in weekdays:
            for time_slot in time_slots:
                rows.append(f"{day} {time_slot}")
        
        # Also add monthly day rows
        for time_slot in time_slots:
            rows.append(f"MONTHLY {time_slot}")
        
        # Initialize matrix with empty strings
        matrix = pd.DataFrame('', index=rows, columns=all_categories)
        matrix.index.name = 'Schedule'
        
        # Fill matrix with device group names
        for _, row in data_df.iterrows():
            category = row['Category']
            device_group = row['Device Group Name']
            recurrence = row['Recurrence']
            weekdays_str = row['Weekdays']
            time_str = row['Start Time']
            
            # Skip if no time or invalid
            if not time_str or time_str == 'N/A':
                continue
            
            # Parse time to nearest 30-minute slot
            try:
                time_parts = time_str.replace('AM', '').replace('PM', '').strip().split(':')
                hour = int(time_parts[0])
                minute = int(time_parts[1])
                
                # Adjust for PM
                if 'PM' in time_str and hour != 12:
                    hour += 12
                elif 'AM' in time_str and hour == 12:
                    hour = 0
                
                # Round to nearest 30-minute slot
                if minute < 15:
                    minute = 0
                elif minute < 45:
                    minute = 30
                else:
                    minute = 0
                    hour = (hour + 1) % 24
                
                time_slot = f"{hour:02d}:{minute:02d}"
                
                # Populate matrix based on recurrence type
                if recurrence == 'MONTHLY':
                    row_name = f"MONTHLY {time_slot}"
                    if row_name in matrix.index and category in matrix.columns:
                        existing = matrix.loc[row_name, category]
                        matrix.loc[row_name, category] = f"{existing}\n{device_group}" if existing else device_group
                elif recurrence == 'WEEKLY' and weekdays_str and weekdays_str != 'N/A':
                    days = [d.strip().upper() for d in weekdays_str.split(',')]
                    for day in days:
                        if len(day) > 3:
                            day = day[:3]
                        row_name = f"{day} {time_slot}"
                        if row_name in matrix.index and category in matrix.columns:
                            existing = matrix.loc[row_name, category]
                            matrix.loc[row_name, category] = f"{existing}\n{device_group}" if existing else device_group
                elif recurrence == 'DAILY':
                    for day in weekdays:
                        row_name = f"{day} {time_slot}"
                        if row_name in matrix.index and category in matrix.columns:
                            existing = matrix.loc[row_name, category]
                            matrix.loc[row_name, category] = f"{existing}\n{device_group}" if existing else device_group
            except Exception:
                continue
        
        return matrix
    
    # Helper function to create schedule matrix view
    def create_matrix_view(data_df: pd.DataFrame, include_tower: bool = False, 
                          all_device_groups: list = None, dg_to_tower_map: dict = None) -> tuple:
        """Create a matrix view with device groups as rows and categories as columns"""
        if data_df.empty and not all_device_groups:
            return pd.DataFrame(), {}
        
        # Add category column
        data_df = data_df.copy()
        if not data_df.empty:
            data_df['Category'] = data_df.apply(lambda row: extract_category(row['Report Name'], row.get('Tower', '')), axis=1)
        
        # Get unique device groups
        if all_device_groups:
            device_groups = sorted(all_device_groups)
        else:
            device_groups = sorted(data_df['Device Group Name'].unique())
        
        # Create device group to tower mapping
        dg_to_tower = {}
        if include_tower:
            if dg_to_tower_map:
                dg_to_tower = dg_to_tower_map
            else:
                for dg in device_groups:
                    dg_data = data_df[data_df['Device Group Name'] == dg]
                    if not dg_data.empty:
                        tower = dg_data['Tower'].iloc[0]
                        dg_to_tower[dg] = tower
                    else:
                        dg_to_tower[dg] = 'N/A'
        
        # Get categories from data
        categories = sorted(data_df['Category'].unique()) if not data_df.empty else []
        
        # Initialize matrix with empty strings
        matrix = pd.DataFrame('', index=device_groups, columns=categories)
        matrix.index.name = 'Device Group'
        
        # Track schedules for conflict detection
        schedule_tracker = {dg: {} for dg in device_groups}
        
        # Fill in the matrix with schedule information
        for _, row in data_df.iterrows():
            dg = row['Device Group Name']
            cat = row['Category']
            recurrence = row['Recurrence']
            report_id = row.get('Report ID', '')
            
            if dg not in device_groups:
                continue
            
            # Helper function to format day with ordinal suffix
            def format_day_ordinal(day):
                """Convert day number to ordinal string (1st, 2nd, 3rd, etc.)"""
                if not day or day == 'N/A':
                    return ''
                try:
                    day_num = int(day)
                    if 10 <= day_num % 100 <= 20:
                        suffix = 'th'
                    else:
                        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day_num % 10, 'th')
                    return f"{day_num}{suffix}"
                except:
                    return str(day)
            
            # Determine the day/weekday based on recurrence type
            if recurrence == 'MONTHLY':
                day_of_month_str = row['Day of Month']
                if day_of_month_str and day_of_month_str != 'N/A':
                    days_list = [d.strip() for d in str(day_of_month_str).split(',')]
                    formatted_days = [format_day_ordinal(d) + ' of Month' for d in days_list if d]
                    day_info = ', '.join(formatted_days)
                else:
                    day_info = ''
            else:
                day_info = row['Weekdays'] if row['Weekdays'] != 'N/A' else ''
            
            # Get time
            time_local = row['Start Time']
            
            # Format cell content
            if recurrence == 'ONCHANGE':
                content = 'YES'
            elif day_info:
                content = f"{day_info} {time_local}" if time_local != 'N/A' else day_info
            else:
                content = time_local if time_local != 'N/A' else ''
            
            # Track for conflict detection
            schedule_key = (day_info, time_local)
            if schedule_key not in schedule_tracker[dg]:
                schedule_tracker[dg][schedule_key] = []
            schedule_tracker[dg][schedule_key].append((cat, report_id))
            
            # Add to matrix
            existing = matrix.loc[dg, cat]
            matrix.loc[dg, cat] = f"{existing}\n{content}".strip() if existing else content
        
        # Add Tower column if requested
        if include_tower:
            matrix = matrix.reset_index()
            tower_values = [dg_to_tower[dg] for dg in matrix['Device Group']]
            matrix.insert(0, 'Tower', tower_values)
        
        return matrix, schedule_tracker
    
    # Helper function to apply conflict highlighting
    def apply_conflict_highlighting(worksheet, matrix, schedule_tracker, start_row=2, has_tower_col=False):
        """Apply highlighting to Excel: Different colors for duplicates, red for missing schedules"""
        from openpyxl.styles import PatternFill, Font
        
        # Define color palette for duplicate schedules
        color_palette = [
            'FFFF00', 'FFA500', 'FF69B4', '00FFFF', 'FF00FF', '90EE90',
            'FFB6C1', 'DDA0DD', 'F0E68C', 'FFE4B5', 'AFEEEE', 'DB7093',
        ]
        
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        black_font = Font(color='000000', bold=True)
        
        # Map device group names to row indices
        if has_tower_col and 'Device Group' in matrix.columns:
            dg_to_row = {matrix.iloc[idx, 1]: idx + start_row for idx in range(len(matrix))}
        else:
            dg_to_row = {dg: idx + start_row for idx, dg in enumerate(matrix.index)}
        
        # Map category to column indices
        col_map = {}
        start_col = 1 if has_tower_col else 2
        for col_idx, col in enumerate(matrix.columns, start=start_col):
            if col in ['Tower', 'Device Group']:
                continue
            col_map[col] = col_idx
        
        # Build schedule text map for duplicate detection
        schedule_text_map = {}
        for dg in (matrix['Device Group'] if has_tower_col else matrix.index):
            for cat in col_map.keys():
                if has_tower_col:
                    cell_value = matrix.loc[matrix['Device Group'] == dg, cat].iloc[0] if len(matrix[matrix['Device Group'] == dg]) > 0 else ''
                else:
                    cell_value = matrix.loc[dg, cat]
                
                if cell_value and cell_value.strip():
                    for schedule_line in cell_value.split('\n'):
                        schedule_line = schedule_line.strip()
                        if schedule_line:
                            if schedule_line not in schedule_text_map:
                                schedule_text_map[schedule_line] = []
                            schedule_text_map[schedule_line].append((dg, cat))
        
        # Assign colors to duplicate schedules
        schedule_to_color = {}
        color_index = 0
        for schedule_text, locations in schedule_text_map.items():
            if len(locations) > 1:
                schedule_to_color[schedule_text] = color_palette[color_index % len(color_palette)]
                color_index += 1
        
        # Create lookup for duplicate schedule colors
        duplicate_schedule_colors = {}
        for schedule_text, locations in schedule_text_map.items():
            if schedule_text in schedule_to_color:
                for dg, cat in locations:
                    duplicate_schedule_colors[(dg, cat, schedule_text)] = schedule_to_color[schedule_text]
        
        # Identify device groups with schedules
        dg_with_schedules = set()
        for dg, schedules in schedule_tracker.items():
            if schedules and any(len(reports) > 0 for reports in schedules.values()):
                dg_with_schedules.add(dg)
        
        # Identify categories with reports
        categories_with_reports = set()
        for dg, schedules in schedule_tracker.items():
            for (day, time), reports in schedules.items():
                if reports:
                    for cat, report_id in reports:
                        categories_with_reports.add(cat)
        
        # Apply highlighting
        for dg in dg_to_row.keys():
            row_idx = dg_to_row[dg]
            for cat, col_idx in col_map.items():
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value) if cell.value else ''
                
                # Check for duplicate schedule
                duplicate_color = None
                if cell_value.strip():
                    for schedule_line in cell_value.split('\n'):
                        schedule_line = schedule_line.strip()
                        if (dg, cat, schedule_line) in duplicate_schedule_colors:
                            duplicate_color = duplicate_schedule_colors[(dg, cat, schedule_line)]
                            break
                
                if duplicate_color:
                    cell.fill = PatternFill(start_color=duplicate_color, end_color=duplicate_color, fill_type='solid')
                    cell.font = black_font
                elif not cell_value.strip() and dg in dg_with_schedules and cat in categories_with_reports:
                    # Red for missing schedules
                    cell.fill = red_fill
                    cell.font = white_font
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(schedule_data)
    
    # Sort by Tower, Device Group Name, and Report Name
    df = df.sort_values(['Tower', 'Device Group Name', 'Report Name'])
    
    # Prepare list of ALL device groups (excluding generic ones)
    all_dg_names = [dg.get('name') for dg in device_groups 
                    if dg.get('name') 
                    and dg.get('name') != 'All Devices' 
                    and not dg.get('name').startswith('DG-')]
    
    # Create mapping of device group name to tower for ALL device groups
    all_dg_to_tower = {}
    for dg_name in all_dg_names:
        dg_data = df[df['Device Group Name'] == dg_name]
        if not dg_data.empty:
            all_dg_to_tower[dg_name] = dg_data['Tower'].iloc[0]
        else:
            all_dg_to_tower[dg_name] = infer_tower(dg_name, '')
    
    # Create output filename with date and time
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_excel = f'{args.prefix}_schedules_{timestamp}.xlsx'
    
    # Add Category column to main dataframe
    df['Category'] = df.apply(lambda row: extract_category(row['Report Name'], row['Tower']), axis=1)
    
    # Get list of unique towers
    towers = sorted(df['Tower'].unique())
    
    # Create Excel writer with multiple sheets
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # Main sheet with all data
        df.to_excel(writer, index=False, sheet_name='All Schedules')
        
        # Create matrix view for all schedules
        matrix_all, tracker_all = create_matrix_view(df, include_tower=True, 
                                                     all_device_groups=all_dg_names, 
                                                     dg_to_tower_map=all_dg_to_tower)
        if not matrix_all.empty:
            matrix_all.to_excel(writer, sheet_name='All Schedules Matrix', index=False)
            ws = writer.sheets['All Schedules Matrix']
            apply_conflict_highlighting(ws, matrix_all, tracker_all, start_row=2, has_tower_col=True)
        
        # Create separate sheets for each tower
        for tower in towers:
            tower_df = df[df['Tower'] == tower].copy()
            tower_dg_names = [dg for dg, t in all_dg_to_tower.items() if t == tower]
            
            if tower_dg_names:
                # Detail sheet
                if not tower_df.empty:
                    tower_df.to_excel(writer, index=False, sheet_name=tower)
                
                # Matrix view sheet
                tower_dg_to_tower = {dg: tower for dg in tower_dg_names}
                matrix_tower, tracker_tower = create_matrix_view(tower_df, include_tower=False, 
                                                                 all_device_groups=tower_dg_names, 
                                                                 dg_to_tower_map=tower_dg_to_tower)
                if not matrix_tower.empty:
                    matrix_tower.to_excel(writer, sheet_name=f'{tower} Matrix')
                    ws = writer.sheets[f'{tower} Matrix']
                    apply_conflict_highlighting(ws, matrix_tower, tracker_tower, start_row=2, has_tower_col=False)
        
        # Create Time Slot Analysis Sheet
        timeslot_df = create_timeslot_analysis(df)
        if not timeslot_df.empty:
            timeslot_df.to_excel(writer, sheet_name='Time Slot Analysis', index=True)
    
    print(f"\n{'='*60}")
    print(f"SUCCESS: Created {output_excel}")
    print(f"{'='*60}")
    print(f"Total report schedules: {len(schedule_data)}")
    print(f"Total device groups (from API): {len(all_dg_names)}")
    print(f"Device groups with schedules: {df['Device Group Name'].nunique()}")
    print(f"Device groups without schedules: {len(all_dg_names) - df['Device Group Name'].nunique()}")
    
    enabled_count = (df['Schedule Enabled'] == 'Yes').sum()
    disabled_count = (df['Schedule Enabled'] == 'No').sum()
    print(f"Enabled schedules: {enabled_count}")
    print(f"Disabled schedules: {disabled_count}")
    
    # Show tower breakdown
    print(f"\nTower Distribution:")
    tower_counts = df['Tower'].value_counts()
    for tower in towers:
        count = tower_counts.get(tower, 0)
        tower_dg_total = sum(1 for dg, t in all_dg_to_tower.items() if t == tower)
        tower_dg_with_schedules = df[df['Tower'] == tower]['Device Group Name'].nunique() if count > 0 else 0
        print(f"  {tower}: {count} reports across {tower_dg_with_schedules}/{tower_dg_total} device groups")
    
    # Show device group breakdown by tower
    print(f"\nTop Device Groups by Tower:")
    for tower in towers:
        tower_df = df[df['Tower'] == tower]
        if not tower_df.empty:
            print(f"\n  {tower}:")
            dg_counts = tower_df['Device Group Name'].value_counts()
            for dg, count in dg_counts.head(5).items():
                print(f"    {dg}: {count} reports")
            if len(dg_counts) > 5:
                print(f"    ... and {len(dg_counts) - 5} more device groups")


if __name__ == '__main__':
    main()
